# -*- coding: utf-8 -*-

__author__ = 'Manfred'
from openpyxl import Workbook, load_workbook # import basic Class, function to read/write xlsx.
from openpyxl.styles import Alignment # import Class for xlsx styles
from dbtest.views import class_info_query, temp_table_update
import random

DEFAULT_SCORE_START_ROW = 6


class XlsxInfo:
    """
    XlsxInfo class is used to store information in xlsx file used by score system 
    """
    def __init__(self, _class_id=None, _scores=None):
        # each teacher teaching a course has a unique class_id
        self.class_id = _class_id
        # all the scores for this class
        self.scores = _scores
        self.error_code = 0
        self.warning_code = 0


# parse the 'xlsx_filename' file in the format of XlsxInfo
def parse_xlsx(xlsx_filename):
    # init the existing xlsx file with read-only mode to workbook(wb)
    wb = load_workbook(filename=xlsx_filename, read_only=True)
    # get the active worksheet(ws) from xlsx workbook
    ws = wb.active

    xlsx_info = XlsxInfo()

    # check whether the syntax remains the same from demo.xlsx
    if ws['A5'].value == '学号':
        # and str(ws['A6'].value).startswith('31')
        if ws['F3'].value is not None:
            class_id = ws['F3'].value
        else:
            xlsx_info.error_str = '[格式错误] 课程编号未填写!'
            xlsx_info.error_code = 3
            return xlsx_info
        print('ok!')
    else:
        xlsx_info.error_str = '[格式错误] 未知格式错误，您可能更改了模板xlsx文件的文本格式!'
        xlsx_info.error_code = 3
        return xlsx_info

    # the default start row for score information
    start_row = DEFAULT_SCORE_START_ROW
    # get the end row of score information
    end_row = ws.get_highest_row()
    print('highest row {}'.format(end_row))
    score_info = []
    for row in range(start_row, end_row + 1):
        print('now processing row {}.'.format(row))
        # if studentID is not specified then skip it, not care about studentName
        if ws['A{}'.format(row)].value is None :
            xlsx_info.error_code = 1
            xlsx_info.error_str = '第{}行: 学生学号不正确.'.format(row)
            print(xlsx_info.error_str)
            continue 

        # if score is not specified then skip it.
        if ws['C{}'.format(row)].value is None :
            xlsx_info.warning_code = 2
            print('row {} in xlsx file: student {}\'s score not specified.'\
                    .format(row, ws['A{}'.format(row)].value))
            continue

        try:
            row_info = {'studentID': ws['A{}'.format(row)].value, 'score': float(ws['C{}'.format(row)].value)}
        except ValueError:
            xlsx_info.error_code = 2    
            # xlsx_info.error_str = 'row {}: score is not a float number.'.format(row)
            xlsx_info.error_str = '第{}行: 成绩不是一个有效的数字.'.format(row)
            print(xlsx_info.error_str) 
            continue

        print(row_info)
        score_info.append(row_info)

    xlsx_info.class_id = class_id 
    xlsx_info.scores = score_info

#    print('after parse')
#    print(xlsx_info.class_id)
#    print(xlsx_info.score_info)
    return xlsx_info

import os
def get_demo_xlsx(course_id):
    wb = Workbook()
    ws = wb.active
    ws.title = 'score sheet 1'

    student_info = class_info_query(course_id)


    # Formatting
    ws.append([])
    ws.append([])
    ws.merge_cells('A1:F2')
    center_align = Alignment(horizontal='center', vertical='center')

    # Initialization
    ws['A1'] = '浙江大学课程考试成绩记录表'
    ws['A1'].alignment = center_align

    ws.append(['教学班名称', '', '教师姓名', '', '课程编号'])
    ws.append([])
    ws.append(['学号', '姓名', '成绩', '备注'])


    # Filling the data.
    ws['F3'] = course_id

    for row in student_info:
        # for row in range(5, 5 + 10):
        # ws['A' + str(row)] = '312xxxxxxx'
        # ws['B' + str(row)] = 'Elder Wang'
        # ws['C' + str(row)] = '{0:.2f}'.format(random.Random().random() * 100)
        if row['score'] is None:
            row_info = [row['studentID'], row['studentName']]
        else:
            row_info = [row['studentID'], row['studentName'], row['score']]
        ws.append(row_info)

    # Save the file
    ws.page_setup.fitToWidth = 1

    # make sure the download dir exists
    download_dir = './download/'
    if not os.path.exists(download_dir):
        os.mkdir(download_dir)

    # and save the file for download
    wb.save(download_dir + "demo.xlsx")


def update_score(xlsx_filename, c_id):
    """
    use the file uploaded by the user to update database
    """

    # parse the xlsx file
    xlsx_info = parse_xlsx(xlsx_filename)

    print('cid')
    print(c_id)
    print("xlsx_info.class_id")
    print(xlsx_info.class_id)

    if xlsx_info.error_code is not 0:
        return xlsx_info.error_str

    # check if the class_id in xlsx file match (front-end)selected c_id 
    if c_id == xlsx_info.class_id:
        return temp_table_update(xlsx_info.class_id, xlsx_info.scores)
    else:
        # return "The class id mentioned in the uploaded file doesn't match your\
 # selected class. Update failed."
        return '文件中的课程编号与选择的不符，上传失败。'


